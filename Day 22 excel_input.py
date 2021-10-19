# -*- coding: utf-8 -*-
"""
Created on Thu Oct 14 22:18:48 2021

@author: 皮皮卡卡
"""
import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb.active
sheet.title = 'just testing'
wb.save('just testing.xlsx') 

wb.create_sheet(index=2, title='Midddle Sheet')

sheet = wb['just testing']
sheet['A1'] = 'hello, world'
print(sheet['A1'].value)

 