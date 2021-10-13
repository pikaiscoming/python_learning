# -*- coding: utf-8 -*-
"""
Created on Tue Oct 12 20:50:59 2021

@author: 皮皮卡卡
"""
import openpyxl as pyxl
from openpyxl.utils import get_column_letter, column_index_from_string #translate column letter or number 

wb = pyxl.load_workbook('example.xlsx') #到這邊都必要的
sheet = wb['sheet1']    #get a cell form the sheet
sheet = wb.active #the sheet which you first get when you open excel.
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_column)
'''
we can get letter of colume and number translating
'''

print(get_column_letter(900))
print(column_index_from_string('A'))
'''
we can use the 'slice' to get specific data from a rectangle range in excel.
'''

for rowcellobject in sheet['B1':'B7']:
    for obj in rowcellobject:
        print(obj.coordinate, obj.value) #是一個正方形，從A1到C3，當然也可以是一格的長方形(單行)
        
        
        
    