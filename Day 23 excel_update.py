# -*- coding: utf-8 -*-
"""
Created on Tue Oct 19 21:36:21 2021

@author: 皮皮卡卡
"""
import openpyxl
from openpyxl.styles import Font #控制字型


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

priceupdates = {'Garlic':3.07,
                'Celery':1.19,
                'Lemon':1.27}


#loop to check 

for rownum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rownum, column=1).value
    
    if produceName in priceupdates: #it is ok to check all of key in dict by for looping.
        sheet.cell(row=rownum, column=2).value = priceupdates[produceName]

word = Font(size=24, name='Times New Roman',italic=True)
'''
name = 字型
bold = 粗體
italic = 斜體
'''
sheet['A1'].Font = word #setup word styles
sheet['A1']= 'hello' #decide the sentence.

        
wb.save('updateproducesales.xlsx')
        
        